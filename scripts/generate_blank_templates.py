#!/usr/bin/env python3
"""Generates blank PT and EN copies of the source workbook: same fixtures,
weights, formulas, and conditional formatting, but with the 5 subjective
criteria columns emptied out so someone else can fill in their own ratings.

Usage:
    .venv/bin/python scripts/generate_blank_templates.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "Notas_da_Copa_2026.xlsx"
OUT_DIR = ROOT / "templates_blank"

CRITERIA_COLUMNS = ("F", "G", "H", "I", "J")
PHASE_SHEETS = (
    "Fase de grupos",
    "16 avos",
    "Oitavas de final",
    "Quartas de final",
    "Semifinais",
    "3º lugar",
    "Final",
)

SHEET_NAME_MAP = {
    "Resumo": "Summary",
    "Quesitos": "Criteria",
    "Fase de grupos": "Group stage",
    "16 avos": "Round of 32",
    "Oitavas de final": "Round of 16",
    "Quartas de final": "Quarterfinals",
    "Semifinais": "Semifinals",
    "3º lugar": "Third place",
    "Final": "Final",
}

# Literal (non-formula) text that needs translating wherever it appears as a
# cell value: sheet titles in A1, the Resumo phase list, headers, row labels.
TEXT_MAP = {
    "Controle de notas — Copa do Mundo 2026": "Match ratings — 2026 World Cup",
    "Indicador": "Indicator",
    "Valor": "Value",
    "Fase": "Phase",
    "Partidas": "Matches",
    "Média das notas": "Average score",
    "Total de partidas": "Total matches",
    "Partidas avaliadas": "Rated matches",
    "Média geral": "Overall average",
    "Maior nota": "Highest score",
    "Menor nota": "Lowest score",
    "Última atualização da estrutura": "Structure last updated",
    "Quesitos e pesos da nota final": "Criteria and weights for the final score",
    "Quesito": "Criterion",
    "Peso": "Weight",
    "Peso (%)": "Weight (%)",
    "Orientação": "Guidance",
    "Total": "Total",
    "1º tempo": "First half",
    "2º tempo": "Second half",
    "Lá e cá": "Back and forth",
    "Emoção": "Emotion",
    "Componente histórico": "Historic component",
    "Nota de 0 a 10 para a qualidade do primeiro tempo.": "Score from 0 to 10 for the quality of the first half.",
    "Nota de 0 a 10 para a qualidade do segundo tempo.": "Score from 0 to 10 for the quality of the second half.",
    "Alternância, ataques dos dois lados, ritmo e sensação de jogo aberto.": (
        "Back-and-forth play, attacks from both sides, pace, and a sense of an open game."
    ),
    "Tensão, imprevisibilidade, viradas, gols tardios e drama.": (
        "Tension, unpredictability, comebacks, late goals, and drama."
    ),
    "Contexto histórico, zebras, recordes, rivalidade e relevância para a Copa.": (
        "Historical context, upsets, records, rivalry, and relevance to the World Cup."
    ),
    "Legenda de cores": "Color legend",
    "Faixa": "Range",
    "Cor": "Color",
    "Uso": "Usage",
    "Aplicada aos quesitos e à nota final": "Applied to each criterion and to the final score",
    "Nota": "Score",
    "Número da partida": "Match number",
    "Data": "Date",
    "Hora (BRT)": "Time (BRT)",
    "Partida": "Match",
    "Placar": "Score",
    "Nota final": "Final score",
    "32 avos": "Round of 32",
}
# The Resumo/Summary sheet also repeats each phase name as plain text (column
# D), not just inside formulas — reuse the same mapping for those cells.
TEXT_MAP.update({k: v for k, v in SHEET_NAME_MAP.items() if k != v})


def sheet_ref(name: str) -> str:
    return f"'{name}'!" if " " in name else f"{name}!"


def clear_criteria_scores(wb: openpyxl.Workbook) -> None:
    for phase in PHASE_SHEETS:
        ws = wb[phase]
        for row in range(3, ws.max_row + 1):
            if ws[f"A{row}"].value is None:
                continue
            for col in CRITERIA_COLUMNS:
                ws[f"{col}{row}"].value = None


def translate_to_english(wb: openpyxl.Workbook) -> None:
    # 1. Rewrite formulas that reference sheet names, while titles are still PT.
    for old, new in SHEET_NAME_MAP.items():
        if old == new:
            continue
        old_ref, new_ref = sheet_ref(old), sheet_ref(new)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("=") and old_ref in cell.value:
                        cell.value = cell.value.replace(old_ref, new_ref)

    # 2. Translate literal text values (titles, headers, labels, descriptions).
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in TEXT_MAP:
                    cell.value = TEXT_MAP[cell.value]

    # 3. Decimal separator in the color-band legend text (Quesitos!B12:B16):
    #    "2–4,99" -> "2-4.99".
    criteria_ws = wb["Quesitos"]
    for row in range(12, 17):
        cell = criteria_ws[f"B{row}"]
        if isinstance(cell.value, str):
            cell.value = cell.value.replace(",", ".").replace("–", "-")

    # 4. Finally, rename the sheet tabs themselves.
    for old, new in SHEET_NAME_MAP.items():
        if old != new:
            wb[old].title = new


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)

    pt_path = OUT_DIR / "Notas_da_Copa_2026_modelo_PT.xlsx"
    wb_pt = openpyxl.load_workbook(SOURCE)
    clear_criteria_scores(wb_pt)
    wb_pt.save(pt_path)
    print(f"wrote {pt_path}")

    en_path = OUT_DIR / "World_Cup_2026_match_ratings_template_EN.xlsx"
    wb_en = openpyxl.load_workbook(SOURCE)
    clear_criteria_scores(wb_en)
    translate_to_english(wb_en)
    wb_en.save(en_path)
    print(f"wrote {en_path}")


if __name__ == "__main__":
    main()
