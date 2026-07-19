"""Reads the "Notas da Copa" workbook and turns it into the plain structures
in models.py. This is the only place that knows about sheet names, cell
layout, or openpyxl — everything downstream (aggregate.py, the API) is
spreadsheet-agnostic.

Deliberately does NOT recompute the "Nota final" column: it reads the value
Excel/LibreOffice/Google Sheets already cached for that formula. The
workbook is the source of truth for that number.
"""
from __future__ import annotations

import re
import warnings
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .models import CRITERIA, CRITERIA_LABELS_PT, ColorBand, CriteriaWeight, Match

PHASE_SHEETS = (
    "Fase de grupos",
    "16 avos",
    "Oitavas de final",
    "Quartas de final",
    "Semifinais",
    "3º lugar",
    "Final",
)

_LABEL_TO_CRITERION = {label: key for key, label in CRITERIA_LABELS_PT.items()}

# Column layout shared by every phase sheet, header row 2, data from row 3.
_COL_MATCH_NUMBER = "A"
_COL_DATE = "B"
_COL_TIME = "C"
_COL_MATCH = "D"
_COL_SCORE = "E"
_COL_CRITERIA_START = "F"  # F..J, in CRITERIA order
_COL_FINAL_SCORE = "K"

_CF_FORMULA_RE = re.compile(
    r"(?P<col>[A-Z]+)\d+\s*(?P<op1>>=|>)\s*(?P<lo>[\d.]+).*?"
    r"(?P<col2>[A-Z]+)\d+\s*(?P<op2><=|<)\s*(?P<hi>[\d.]+)"
)


class WorkbookFormatError(ValueError):
    """Raised when the workbook doesn't match the expected "Notas da Copa" layout."""


def load_criteria_weights(wb: openpyxl.Workbook) -> list[CriteriaWeight]:
    ws = wb["Quesitos"]
    weights: list[CriteriaWeight] = []
    for row in range(3, 8):
        label = ws.cell(row=row, column=1).value
        weight = ws.cell(row=row, column=2).value
        criterion = _LABEL_TO_CRITERION.get(str(label).strip() if label else "")
        if criterion is None or weight is None:
            raise WorkbookFormatError(
                f"Quesitos!A{row} = {label!r} não corresponde a um quesito esperado "
                f"({', '.join(CRITERIA_LABELS_PT.values())})."
            )
        weights.append(CriteriaWeight(criterion=criterion, weight=float(weight)))
    return weights


def load_color_bands(wb: openpyxl.Workbook) -> list[ColorBand]:
    """Reads the 5 conditional-formatting rules applied to the "Nota final"
    columns (any phase sheet has them; they're kept in sync by hand in the
    workbook). Falls back to the documented defaults if parsing fails, e.g.
    because a third party stripped conditional formatting when re-saving.
    """
    ws = wb[PHASE_SHEETS[0]]
    bands: list[ColorBand] = []
    try:
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.type != "expression" or not rule.formula:
                    continue
                m = _CF_FORMULA_RE.search(rule.formula[0])
                if not m:
                    continue
                lo = float(m.group("lo"))
                hi = float(m.group("hi"))
                inclusive = m.group("op2") == "<="
                fg = _argb_to_hex(rule.dxf.font.color.rgb) if rule.dxf.font else "#000000"
                bg = (
                    _argb_to_hex(rule.dxf.fill.fgColor.rgb)
                    if rule.dxf.fill and rule.dxf.fill.fgColor
                    else "#FFFFFF"
                )
                bands.append(ColorBand(lower=lo, upper=hi, upper_inclusive=inclusive, fg=fg, bg=bg))
    except Exception:  # pragma: no cover - defensive, see docstring
        bands = []
    if len(bands) != 5:
        from .models import DEFAULT_COLOR_BANDS

        warnings.warn(
            "Não foi possível ler as faixas de cor da planilha; usando as faixas padrão.",
            stacklevel=2,
        )
        return list(DEFAULT_COLOR_BANDS)
    return sorted(bands, key=lambda b: b.lower)


def _argb_to_hex(argb: str | None) -> str:
    if not argb:
        return "#000000"
    return "#" + argb[-6:]


def _parse_score(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load_matches(wb: openpyxl.Workbook) -> list[Match]:
    matches: list[Match] = []
    for phase in PHASE_SHEETS:
        ws: Worksheet = wb[phase]
        for row in range(3, ws.max_row + 1):
            match_number = ws[f"{_COL_MATCH_NUMBER}{row}"].value
            if match_number is None:
                continue
            match_label = ws[f"{_COL_MATCH}{row}"].value
            if not match_label or "x" not in str(match_label):
                continue
            home, _, away = str(match_label).partition(" x ")

            scores = {}
            for offset, criterion in enumerate(CRITERIA):
                col = chr(ord(_COL_CRITERIA_START) + offset)
                scores[criterion] = _parse_score(ws[f"{col}{row}"].value)

            match_date_value = ws[f"{_COL_DATE}{row}"].value
            match_date = match_date_value.date() if hasattr(match_date_value, "date") else None

            # Excel/Sheets sometimes autocorrects a typed score like "2-0"
            # into a date; there's no way to recover the real score from
            # that, so surface it as missing rather than an ISO timestamp.
            score_value = ws[f"{_COL_SCORE}{row}"].value
            score = None if hasattr(score_value, "date") else score_value

            matches.append(
                Match(
                    phase=phase,
                    match_number=int(match_number),
                    match_date=match_date,
                    match_time=ws[f"{_COL_TIME}{row}"].value,
                    home_team=home.strip(),
                    away_team=away.strip(),
                    score=score,
                    scores=scores,
                    final_score=_parse_score(ws[f"{_COL_FINAL_SCORE}{row}"].value),
                )
            )
    return matches


def load_workbook_data(path: str | Path) -> tuple[list[Match], list[CriteriaWeight], list[ColorBand]]:
    """Convenience entrypoint: opens the workbook once and returns everything
    aggregate.py and the API need."""
    wb = openpyxl.load_workbook(path, data_only=True)
    missing = [s for s in (*PHASE_SHEETS, "Quesitos") if s not in wb.sheetnames]
    if missing:
        raise WorkbookFormatError(f"Faltam abas na planilha: {', '.join(missing)}")
    matches = load_matches(wb)
    weights = load_criteria_weights(wb)
    bands = load_color_bands(wb)
    return matches, weights, bands
