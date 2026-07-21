from match_ratings import aggregate as agg
from match_ratings.xlsx_loader import load_workbook_data


def test_summary_mirrors_resumo_sheet(workbook_path):
    """aggregate.summary must reproduce whatever the Resumo sheet's own
    cached formulas currently say — compared live against the sheet rather
    than a frozen snapshot, since the personal workbook keeps changing as
    more matches get rated."""
    import openpyxl

    matches, _weights, _bands = load_workbook_data(workbook_path)
    s = agg.summary(matches)

    resumo = openpyxl.load_workbook(workbook_path, data_only=True)["Resumo"]
    assert s.rated_matches == resumo["B5"].value
    assert s.overall_average == resumo["B6"].value
    assert s.highest_score == resumo["B7"].value
    assert s.lowest_score == resumo["B8"].value


def test_top_matches_is_sorted_descending_and_all_rated(workbook_path):
    matches, _weights, _bands = load_workbook_data(workbook_path)
    top = agg.top_matches(matches, limit=10)

    assert len(top) == 10
    assert all(m.is_rated for m in top)
    scores = [m.final_score for m in top]
    assert scores == sorted(scores, reverse=True)


def test_team_stats_only_counts_matches_the_team_played(workbook_path):
    matches, _weights, _bands = load_workbook_data(workbook_path)
    stats = {t.team: t for t in agg.team_stats(matches)}

    argentina = stats["Argentina"]
    expected_matches = [
        m
        for m in agg.rated(matches)
        if m.home_team == "Argentina" or m.away_team == "Argentina"
    ]
    assert argentina.matches_played == len(expected_matches)
    assert argentina.best_match.final_score >= argentina.worst_match.final_score


def test_criteria_averages_cover_all_five_in_sheet_order(workbook_path):
    matches, weights, _bands = load_workbook_data(workbook_path)
    averages = agg.criteria_averages(matches, weights)

    assert [c.criterion for c in averages] == [
        "first_half",
        "second_half",
        "back_and_forth",
        "emotion",
        "historic_component",
    ]
    assert all(0 <= c.average <= 10 for c in averages)


def test_filter_matches_by_phase_and_team(workbook_path):
    matches, _weights, _bands = load_workbook_data(workbook_path)

    group_stage = agg.filter_matches(matches, phase="Fase de grupos")
    assert all(m.phase == "Fase de grupos" for m in group_stage)

    brazil_matches = agg.filter_matches(matches, team="brasil")
    assert all("brasil" in m.home_team.lower() or "brasil" in m.away_team.lower() for m in brazil_matches)


def test_top_teams_extends_past_limit_on_ties(workbook_path):
    matches, _weights, _bands = load_workbook_data(workbook_path)

    top10 = agg.top_teams(matches, limit=10)
    all_teams = agg.team_stats(matches)
    cutoff = top10[-1].average_final_score
    expected_count = sum(1 for t in all_teams if t.average_final_score >= cutoff)

    assert len(top10) == expected_count
    assert len(top10) >= 10
    scores = [t.average_final_score for t in top10]
    assert scores == sorted(scores, reverse=True)


def test_top_teams_ascending_gives_worst_first(workbook_path):
    matches, _weights, _bands = load_workbook_data(workbook_path)

    worst10 = agg.top_teams(matches, limit=10, ascending=True)
    best10 = agg.top_teams(matches, limit=10, ascending=False)

    assert worst10[0].average_final_score <= worst10[-1].average_final_score
    assert worst10[0].team != best10[0].team


def test_criterion_ranking_extends_past_limit_on_ties(workbook_path):
    matches, _weights, _bands = load_workbook_data(workbook_path)

    top10_emotion = agg.criterion_ranking(matches, "emotion", limit=10)
    assert all(m.scores["emotion"] is not None for m in top10_emotion)
    scores = [m.scores["emotion"] for m in top10_emotion]
    assert scores == sorted(scores, reverse=True)
    # "emotion" has more than 10 matches scoring a 10 (see the earlier
    # perfect-10 test), so the tie-break must have pulled in extras.
    assert len(top10_emotion) > 10

    worst10_first_half = agg.criterion_ranking(matches, "first_half", limit=10, ascending=True)
    scores2 = [m.scores["first_half"] for m in worst10_first_half]
    assert scores2 == sorted(scores2)


def test_filter_matches_by_criterion_score(workbook_path):
    matches, _weights, _bands = load_workbook_data(workbook_path)

    perfect_emotion = agg.filter_matches(matches, criterion="emotion", criterion_min=10)
    assert len(perfect_emotion) > 0
    assert all(m.scores["emotion"] == 10 for m in perfect_emotion)

    low_first_half = agg.filter_matches(matches, criterion="first_half", criterion_max=3)
    assert all(m.scores["first_half"] is not None and m.scores["first_half"] <= 3 for m in low_first_half)
